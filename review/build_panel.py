#!/usr/bin/env python3
"""
Rebuild the trust-month analysis panel directly from NHS England's published
Diagnostic Imaging Dataset tables in datain/raw_trust/.

Why this exists
---------------
The derived per-financial-year .dta files this project previously used were
deleted from the repository, so nothing could be reproduced from source. They
were also truncated: the 2023-24 extract was pulled before the year was
complete, so October 2023 onward was progressively under-counted and February
and March 2024 were recorded as zeros. Rebuilding from the published tables
fixes both problems and extends the series to March 2025.

Provenance
----------
Everything used here comes from two tables, published per financial year:

  Table 4  Count of imaging activity, using groups of tests suitable for
           diagnosing cancer, labelled by body site  -> event counts
  Table 5  The same groups, median days from request to test   -> waiting times

Both are provider x test-group x source-setting x month. Source setting takes
two values, "All" and "GP Direct Access", which is what makes the
difference-in-differences possible. The five test groups map onto the policy as:

  Chest (X-ray)                        covered
  Chest and/or abdomen (CT)            covered
  Brain (MRI)                          covered
  Abdomen and/or pelvis (Ultrasound)   covered
  Kidney or Bladder (Ultrasound)       NOT named in the guidance -> comparator

Format drift
------------
The header row is not in a fixed position: Table 4 uses row 13 throughout, but
Table 5 uses row 14 in every year except 2022-23, where it uses row 13. The
2024-25 folder is also named with a hyphen rather than an underscore. Header
rows and columns are therefore detected by content, not by position.

Usage
    python review/build_panel.py [--raw datain/raw_trust] [--out review/results]
"""
import argparse
import glob
import os
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

MONTHS = {"Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9,
          "Oct": 10, "Nov": 11, "Dec": 12, "Jan": 1, "Feb": 2, "Mar": 3}

COVERED = ["Chest (X-ray)", "Chest and/or abdomen (CT)", "Brain (MRI)",
           "Abdomen and/or pelvis (Ultrasound)"]
COMPARATOR = "Kidney or Bladder (Ultrasound)"


def read_table(pattern, value_name):
    """Read one DID table across all financial years into long format."""
    frames, meta = [], []
    for path in sorted(glob.glob(pattern)):
        year = os.path.basename(os.path.dirname(path)).replace("-", "_")
        fy_start = int(year.split("_")[0])
        ws = openpyxl.load_workbook(path, read_only=True, data_only=True)["Provider"]
        head = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))

        header_row = None
        for i, row in enumerate(head):
            vals = [str(c) for c in row if c is not None]
            if "Org Code" in vals and any(v in MONTHS for v in vals):
                header_row = i
                break
        if header_row is None:
            raise SystemExit(f"no header row found in {path}")

        hdr = head[header_row]
        col = {str(c): i for i, c in enumerate(hdr) if c is not None}
        month_cols = {i: MONTHS[str(c)] for i, c in enumerate(hdr) if str(c) in MONTHS}
        meta.append((year, header_row + 1, len(month_cols)))

        rows = []
        for r in ws.iter_rows(min_row=header_row + 2, values_only=True):
            org = r[col["Org Code"]]
            if org is None or str(org).strip() in ("-", ""):
                continue           # the "-" rows are the ENGLAND totals
            org = str(org).strip()
            test, src = str(r[col["Test"]]), str(r[col["Source setting"]])
            for i, month in month_cols.items():
                v = pd.to_numeric(r[i], errors="coerce")
                if pd.isna(v):
                    continue
                year_of = fy_start if month >= 4 else fy_start + 1
                rows.append((org, test, src, pd.Timestamp(year_of, month, 1), v))
        frames.append(pd.DataFrame(rows, columns=["orgcode", "test", "src", "ym", value_name]))

    print(f"  {os.path.basename(pattern)}")
    for year, hr, nm in meta:
        print(f"    {year}: header row {hr}, {nm} months")
    return pd.concat(frames, ignore_index=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default="datain/raw_trust")
    ap.add_argument("--out", default="review/results")
    args = ap.parse_args()

    print("Reading published DID tables")
    counts = read_table(os.path.join(args.raw, "*", "DID-Table-4-*.xlsx"), "events")
    waits = read_table(os.path.join(args.raw, "*", "DID-Table-5-*.xlsx"), "wait")

    panel = counts.merge(waits, on=["orgcode", "test", "src", "ym"], how="outer")
    panel = panel[panel.orgcode.str.startswith("R")]      # NHS trusts only

    Path(args.out).mkdir(parents=True, exist_ok=True)
    dest = Path(args.out) / "panel_from_raw.csv"
    panel.sort_values(["orgcode", "test", "src", "ym"]).to_csv(dest, index=False)

    print(f"\nPanel written to {dest}")
    print(f"  rows {len(panel):,}   trusts {panel.orgcode.nunique()}   "
          f"months {panel.ym.nunique()}   "
          f"{panel.ym.min().date()} to {panel.ym.max().date()}")
    print(f"  events present {panel.events.notna().sum():,}   "
          f"waits present {panel.wait.notna().sum():,}")

    print("\nMonthly coverage of the four covered test groups, GP Direct Access:")
    cov = panel[panel.test.isin(COVERED) & (panel.src == "GP Direct Access")]
    by_month = cov.groupby("ym").agg(events=("events", "sum"),
                                     trusts=("orgcode", "nunique"))
    tail = by_month.loc["2023-09-01":]
    for ym, r in tail.iterrows():
        print(f"  {ym.date()}  {r.events:>10,.0f}  {int(r.trusts)} trusts")
    print("\n  February and March 2024 are populated here; they were zeros in the")
    print("  deleted .dta extract, which was pulled before the year was complete.")


if __name__ == "__main__":
    main()
